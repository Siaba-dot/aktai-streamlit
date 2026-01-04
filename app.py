

# -*- coding: utf-8 -*-
"""
Aktų generatorius (Streamlit Cloud-ready), tik XLSX (be PDF).
- Įkeliamas CSV/XLSX su stulpeliais (min): Objekto adresas, Paslaugos pavadinimas, Įkainis, Plotas (m2).
- Papildomi (nebūtina): Skyrius, Užsakovas, Vykdytojas, Sutarties numeris, Vadybininkas.
- UI: vadybininko filtras (pasirenkama), vieno objekto aktas, bendras aktas pagal sutartį.

Svarbu: suderinamumas su openpyxl 3.0.x ir 3.1.x.
"""

import io
import sys
import unicodedata
from typing import List, Dict, Optional, Tuple
from decimal import Decimal, ROUND_HALF_UP

import streamlit as st
import pandas as pd

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, NamedStyle
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName

# ===== Konfigūros =====
PVM_DEFAULT = Decimal("21.00")
FMT_MONEY = "#,##0.00"
FMT_QTY = "#,##0.00"
ROW_TABLE_START = 9
MAX_LINES = 40
# Jei tavo Excel naudoja ';' formulėms – pakeisk į ';'
SEP = ","

SHEET_DB = "KATALOGAS"
SHEET_META = "META"
SHEET_LIST = "LISTOS"
SHEET_MAP = "MAP"
SHEET_AKT = "AKTAS"
SHEET_SUM = "BENDRAS"


# ===== Pagalbinės =====
def strip_accents(s: str) -> str:
    if s is None:
        return ""
    return "".join(ch for ch in unicodedata.normalize("NFD", str(s)) if unicodedata.category(ch) != "Mn")


def norm(s: str) -> str:
    return strip_accents(s).lower().strip()


def detect_delimiter(sample: str) -> Optional[str]:
    c_semi = sample.count(";")
    c_comma = sample.count(",")
    return ";" if c_semi > c_comma else None


def dec2(v) -> float:
    return float(Decimal(str(v).replace(",", ".")).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def create_named_styles(wb: Workbook) -> None:
    if "Money" not in wb.named_styles:
        stl = NamedStyle(name="Money")
        stl.number_format = FMT_MONEY
        stl.alignment = Alignment(horizontal="right")
        wb.add_named_style(stl)
    if "Qty" not in wb.named_styles:
        stl = NamedStyle(name="Qty")
        stl.number_format = FMT_QTY
        stl.alignment = Alignment(horizontal="right")
        wb.add_named_style(stl)


def set_borders(ws, rng: str, thick: bool = False) -> None:
    side = Side(style="thick" if thick else "thin")
    for row in ws[rng]:
        for c in row:
            c.border = Border(left=side, right=side, top=side, bottom=side)


def autosize(ws) -> None:
    for col in ws.columns:
        try:
            letter = col[0].column_letter
        except Exception:
            continue
        max_len = 0
        for cell in col:
            val = str(cell.value) if cell.value is not None else ""
            max_len = max(max_len, len(val))
        ws.column_dimensions[letter].width = min(max_len + 2, 60)


def f_FILTER(range_txt: str, condition_txt: str, ifempty_txt: str) -> str:
    return f"=FILTER({range_txt}{SEP}{condition_txt}{SEP}{ifempty_txt})"


def f_SEARCH(needle: str, haystack: str) -> str:
    return f"ISNUMBER(SEARCH({needle}{SEP}{haystack}))"


def add_defined_name(wb: Workbook, name: str, attr_text: str) -> None:
    """
    Suderinamumo adapteris: openpyxl 3.1.x turi .add(), 3.0.x turėjo .append()
    """
    dn = DefinedName(name=name, attr_text=attr_text)
    container = wb.defined_names
    if hasattr(container, "add"):
        container.add(dn)
    else:
        # senesnė openpyxl atšaka
        container.append(dn)


# ===== Failo skaitymas + stulpelių atpažinimas =====
def read_catalog(uploaded) -> pd.DataFrame:
    """Skaito CSV/XLSX; aptinka skyriklį; normalizuoja stulpelių pavadinimus (be diakritikų)."""
    if uploaded.name.lower().endswith(".csv"):
        head = uploaded.read(4096).decode("utf-8", errors="ignore")
        uploaded.seek(0)
        delim = detect_delimiter(head)
        df = pd.read_csv(uploaded, sep=delim) if delim else pd.read_csv(uploaded)
    else:
        df = pd.read_excel(uploaded, engine="openpyxl")
    # normalizuojame antraštes (tik DataFrame viduje; vartotojui rodom jo originalias)
    df.columns = [norm(c) for c in df.columns]
    return df


def map_columns(df: pd.DataFrame) -> Dict[str, Optional[str]]:
    """
    Map į laukus:
    address, service, rate, qty, skyrius, uzsakovas, vykdytojas, sutartis, vadybininkas
    """
    col_map: Dict[str, Optional[str]] = {}
    for c in df.columns:
        cn = norm(c)
        if "adres" in cn:
            col_map["address"] = c
        elif "paslaug" in cn:
            col_map["service"] = c
        elif "ikain" in cn or "įkain" in cn:
            col_map["rate"] = c
        elif any(k in cn for k in ["kiek", "plot", "m2", "m3", "m³", "vnt", "val", "apimt", "sanaud"]):
            col_map["qty"] = c
        elif "skyri" in cn:
            col_map["skyrius"] = c
        elif "uzsakov" in cn or "užsakov" in cn:
            col_map["uzsakovas"] = c
        elif "vykdytoj" in cn:
            col_map["vykdytojas"] = c
        elif "sutart" in cn:
            col_map["sutartis"] = c
        elif "vadybin" in cn:
            col_map["vadybininkas"] = c

    for req in ("address", "service", "rate", "qty"):
        if req not in col_map:
            raise ValueError(f"Nerastas stulpelis '{req}' (trūksta vieno iš: adresas/paslauga/įkainis/kiekis).")

    for opt in ("skyrius", "uzsakovas", "vykdytojas", "sutartis", "vadybininkas"):
        col_map.setdefault(opt, None)

    return col_map


# ===== Vieno objekto aktas (su dropdown) =====
def build_workbook_single(
    df: pd.DataFrame,
    col_map: Dict[str, Optional[str]],
    selection_mode: str,  # "adresas" arba "skyrius"
    selected_key: Optional[str],
    header_date: str,
    header_year_exec: str,
) -> Workbook:
    wb = Workbook()
    create_named_styles(wb)

    # KATALOGAS
    wsK = wb.active
    wsK.title = SHEET_DB
    wsK.append(["Adresas", "Paslauga", "Įkainis (be PVM)", "Numatytas kiekis",
                "Skyrius", "Užsakovas", "Vykdytojas", "Sutarties nr.", "Vadybininkas"])
    for _, row in df.iterrows():
        wsK.append([
            str(row[col_map["address"]]).strip(),
            str(row[col_map["service"]]).strip(),
            dec2(row[col_map["rate"]]),
            dec2(row[col_map["qty"]]),
            str(row[col_map["skyrius"]]).strip() if col_map["skyrius"] else "",
            str(row[col_map["uzsakovas"]]).strip() if col_map["uzsakovas"] else "",
            str(row[col_map["vykdytojas"]]).strip() if col_map["vykdytojas"] else "",
            str(row[col_map["sutartis"]]).strip() if col_map["sutartis"] else "",
            str(row[col_map["vadybininkas"]]).strip() if col_map["vadybininkas"] else "",
        ])
    autosize(wsK)

    # MAP (unikalus adresas -> meta)
    wsMap = wb.create_sheet(SHEET_MAP)
    wsMap.append(["Adresas", "Skyrius", "Užsakovas", "Vykdytojas", "Sutarties nr."])
    addr_series = df[col_map["address"]].astype(str).str.strip()
    grp = df.assign(_addr=addr_series).groupby("_addr").first()
    for addr, r in grp.iterrows():
        wsMap.append([
            addr,
            (str(r[col_map["skyrius"]]).strip() if col_map["skyrius"] else ""),
            (str(r[col_map["uzsakovas"]]).strip() if col_map["uzsakovas"] else ""),
            (str(r[col_map["vykdytojas"]]).strip() if col_map["vykdytojas"] else ""),
            (str(r[col_map["sutartis"]]).strip() if col_map["sutartis"] else ""),
        ])
    autosize(wsMap)

    # META
    wsM = wb.create_sheet(SHEET_META)
    wsM["A1"], wsM["B1"] = "Užsakovas", ""
    wsM["A2"], wsM["B2"] = "Vykdytojas", ""
    wsM["A3"], wsM["B3"] = "Sutarties nr.", ""
    wsM["A4"], wsM["B4"] = "Objektas / adresas", selected_key if (selection_mode == "adresas" and selected_key) else ""
    wsM["A5"], wsM["B5"] = "Skyrius", selected_key if (selection_mode == "skyrius" and selected_key) else ""
    wsM["A6"], wsM["B6"] = "Akto data", header_date
    wsM["A7"], wsM["B7"] = "Atlikimo metai", header_year_exec
    autosize(wsM)

    # vardiniai pavadinimai (NAUJA: per adapterį)
    add_defined_name(wb, "AdrSelected", f"'{SHEET_META}'!$B$4")
    add_defined_name(wb, "SkySelected", f"'{SHEET_META}'!$B$5")

    # autopildymas iš MAP
    wsM["B1"] = (
        f'=IFERROR(IF(LEN(B4)>0, INDEX(FILTER({SHEET_MAP}!C2:C100000{SEP}{SHEET_MAP}!A2:A100000=B4),1)'
        f'{SEP}IF(LEN(B5)>0, INDEX(FILTER({SHEET_MAP}!C2:C100000{SEP}{SHEET_MAP}!B2:B100000=B5),1){SEP}"" )), "")'
    )
    wsM["B2"] = (
        f'=IFERROR(IF(LEN(B4)>0, INDEX(FILTER({SHEET_MAP}!D2:D100000{SEP}{SHEET_MAP}!A2:A100000=B4),1)'
        f'{SEP}IF(LEN(B5)>0, INDEX(FILTER({SHEET_MAP}!D2:D100000{SEP}{SHEET_MAP}!B2:B100000=B5),1){SEP}"" )), "")'
    )
    wsM["B3"] = (
        f'=IFERROR(IF(LEN(B4)>0, INDEX(FILTER({SHEET_MAP}!E2:E100000{SEP}{SHEET_MAP}!A2:A100000=B4),1)'
        f'{SEP}IF(LEN(B5)>0, INDEX(FILTER({SHEET_MAP}!E2:E100000{SEP}{SHEET_MAP}!B2:B100000=B5),1){SEP}"" )), "")'
    )

    # LISTOS (dropdown šaltiniai)
    wsL = wb.create_sheet(SHEET_LIST)
    # adresai
    wsL["A1"] = "Adresų paieška:"; wsL["B1"] = ""
    wsL["A2"] = f"=UNIQUE({SHEET_DB}!A2:A100000)"
    wsL["C1"] = "Filtruoti adresai (DV)"
    wsL["C2"] = f'=FILTER(A2#{SEP}ISNUMBER(SEARCH(B1{SEP}A2#)){SEP}"Nėra atitikmenų")'
    # skyriai
    wsL["E1"] = "Skyrių paieška:"; wsL["F1"] = ""
    wsL["E2"] = f"=UNIQUE({SHEET_DB}!E2:E100000)"
    wsL["G1"] = "Filtruoti skyriai (DV)"
    wsL["G2"] = f'=FILTER(E2#{SEP}ISNUMBER(SEARCH(F1{SEP}E2#)){SEP}"Nėra")'
    # paslaugos pagal pasirinktą raktą
    wsL["I1"] = "Paslaugų paieška:"; wsL["J1"] = ""
    wsL["I2"] = (
        f'=IF(LEN({SHEET_META}!B4)>0,'
        f'FILTER({SHEET_DB}!B2:B100000{SEP}{SHEET_DB}!A2:A100000={SHEET_META}!B4)'
        f'{SEP}IF(LEN({SHEET_META}!B5)>0,'
        f'FILTER({SHEET_DB}!B2:B100000{SEP}{SHEET_DB}!E2:E100000={SHEET_META}!B5)'
        f'{SEP}"Nėra"))'
    )
    wsL["J2"] = f'=FILTER(I2#{SEP}ISNUMBER(SEARCH(J1{SEP}I2#)){SEP}"Nėra")'

    # vardiniai pavadinimai Listos lape (NAUJA: per adapterį)
    add_defined_name(wb, "AdresaiDV",   f"'{SHEET_LIST}'!$C$2#")
    add_defined_name(wb, "SkyriaiDV",   f"'{SHEET_LIST}'!$G$2#")
    add_defined_name(wb, "PaslaugosDV", f"'{SHEET_LIST}'!$J$2#")

    # AKTAS
    ws = wb.create_sheet(SHEET_AKT)
    labels = ["Užsakovas", "Vykdytojas", "Sutarties nr.", "Objektas / adresas", "Skyrius", "Akto data"]
    for i, lab in enumerate(labels, start=1):
        ws[f"A{i}"] = f'="{lab}: " & {SHEET_META}!B{i}'
    ws["A7"] = f'="Atlikimo data: " & {SHEET_META}!B7 & " m. "'
    ws["A8"], ws["B8"], ws["C8"], ws["D8"], ws["E8"] = \
        "Eil. Nr.", "Paslaugos pavadinimas", "Kiekis", "Įkainis (be PVM)", "Suma (be PVM)"
    set_borders(ws, "A8:E8", thick=True)

    # DV: adresai/skyriai META lape
    dv_addr = DataValidation(type="list", formula1="=AdresaiDV", allow_blank=True)
    dv_sky = DataValidation(type="list", formula1="=SkyriaiDV", allow_blank=True)
    wsM.add_data_validation(dv_addr); dv_addr.add(wsM["B4"])
    wsM.add_data_validation(dv_sky); dv_sky.add(wsM["B5"])

    # DV: paslaugos lentelėje
    dv_service = DataValidation(type="list", formula1="=PaslaugosDV", allow_blank=True)
    ws.add_data_validation(dv_service)
    dv_nonneg = DataValidation(type="decimal", operator="greaterThanOrEqual", formula1="0", allow_blank=True)
    ws.add_data_validation(dv_nonneg)

    for idx in range(MAX_LINES):
        r = ROW_TABLE_START + idx
        ws.cell(r, 1).value = idx + 1
        dv_service.add(ws.cell(r, 2))
        ws.cell(r, 3).value = (
            f'=IFERROR(INDEX(FILTER({SHEET_DB}!D2:D100000'
            f'{SEP}({SHEET_DB}!B2:B100000=B{r})*('
            f'IF(LEN({SHEET_META}!B4)>0,{SHEET_DB}!A2:A100000={SHEET_META}!B4,{SHEET_DB}!E2:E100000={SHEET_META}!B5'
            f'))){SEP}1){SEP}"" )'
        )
        ws.cell(r, 3).number_format = FMT_QTY
        dv_nonneg.add(ws.cell(r, 3))
        ws.cell(r, 4).value = (
            f'=IFERROR(INDEX(FILTER({SHEET_DB}!C2:C100000'
            f'{SEP}({SHEET_DB}!B2:B100000=B{r})*('
            f'IF(LEN({SHEET_META}!B4)>0,{SHEET_DB}!A2:A100000={SHEET_META}!B4,{SHEET_DB}!E2:E100000={SHEET_META}!B5'
            f'))){SEP}1){SEP}"" )'
        )
        ws.cell(r, 4).number_format = FMT_MONEY
        dv_nonneg.add(ws.cell(r, 4))
        ws.cell(r, 5).value = f"=C{r}*D{r}"
        ws.cell(r, 5).number_format = FMT_MONEY

    set_borders(ws, f"A{ROW_TABLE_START}:E{ROW_TABLE_START+MAX_LINES-1}")
    ws["D12"] = "Suma (be PVM):"; ws["E12"] = f"=SUM(E{ROW_TABLE_START}:E{ROW_TABLE_START+MAX_LINES-1})"
    ws["D13"] = f"PVM {float(PVM_DEFAULT)}%:"; ws["E13"] = f"=E12*{float(PVM_DEFAULT)/100}"
    ws["D14"] = "Suma su PVM:"; ws["E14"] = "=E12+E13"

    for s in (ws, wsM, wsL, wsMap, wsK):
        autosize(s)

    return wb


# ===== Bendras aktas pagal sutartį =====
def _write_address_block(
    ws,
    start_row: int,
    address: str,
    meta: Dict[str, str],
    rows: List[Tuple[str, float, float]]
) -> Tuple[int, str]:
    r = start_row
    ws[f"A{r}"] = f"Objekto adresas: {address}"; ws[f"A{r}"].font = Font(bold=True); r += 1
    ws[f"A{r}"] = f"Užsakovas: {meta.get('uzsakovas','')}"; r += 1
    ws[f"A{r}"] = f"Vykdytojas: {meta.get('vykdytojas','')}"; r += 1
    ws[f"A{r}"] = f"Sutarties nr.: {meta.get('sutartis','')}"; r += 1
    if meta.get('skyrius'):
        ws[f"A{r}"] = f"Skyrius: {meta.get('skyrius','')}"; r += 1

    ws[f"A{r}"], ws[f"B{r}"], ws[f"C{r}"], ws[f"D{r}"], ws[f"E{r}"] = \
        "Eil. Nr.", "Paslaugos pavadinimas", "Kiekis", "Įkainis (be PVM)", "Suma (be PVM)"
    set_borders(ws, f"A{r}:E{r}", thick=True)
    r += 1

    first_data_row = r
    for i, (service, qty, rate) in enumerate(rows, start=1):
        ws[f"A{r}"] = i
        ws[f"B{r}"] = service
        ws[f"C{r}"] = qty;  ws[f"C{r}"].number_format = FMT_QTY
        ws[f"D{r}"] = rate; ws[f"D{r}"].number_format = FMT_MONEY
        ws[f"E{r}"] = f"=C{r}*D{r}"; ws[f"E{r}"].number_format = FMT_MONEY
        r += 1

    last_data_row = r - 1
    ws[f"D{r}"] = "Suma (be PVM):"
    if last_data_row >= first_data_row:
        ws[f"E{r}"] = f"=SUM(E{first_data_row}:E{last_data_row})"
    else:
        ws[f"E{r}"] = 0
    ws[f"E{r}"].number_format = FMT_MONEY
    set_borders(ws, f"D{r}:E{r}", thick=True)
    subtotal_ref = f"E{r}"
    r += 1

    ws[f"D{r}"] = f"PVM {float(PVM_DEFAULT)}%:"
    ws[f"E{r}"] = f"={subtotal_ref}*{float(PVM_DEFAULT)/100}"; ws[f"E{r}"].number_format = FMT_MONEY
    r += 1

    ws[f"D{r}"] = "Suma su PVM:"
    ws[f"E{r}"] = f"={subtotal_ref}+E{r-1}"; ws[f"E{r}"].number_format = FMT_MONEY
    set_borders(ws, f"D{r-2}:E{r}", thick=True)
    r += 2

    return r, subtotal_ref


def build_workbook_consolidated(
    df: pd.DataFrame,
    col_map: Dict[str, Optional[str]],
    contract_no: str,
    addresses_selected: List[str],
    header_date: str,
    header_year_exec: str
) -> Workbook:
    if not col_map["sutartis"]:
        raise ValueError("Faile nėra „Sutarties numeris“ – negaliu formuoti bendro akto.")

    dfc = df[df[col_map["sutartis"]].astype(str).str.strip() == contract_no].copy()
    if dfc.empty:
        raise ValueError("Pagal pasirinktą sutartį nerasta įrašų.")

    all_addrs = sorted(set(dfc[col_map["address"]].astype(str).str.strip()))
    if not addresses_selected:
        addresses_selected = all_addrs

    wb = Workbook()
    create_named_styles(wb)

    # META (informacinė)
    wsM = wb.active
    wsM.title = SHEET_META
    wsM["A1"], wsM["B1"] = "Sutarties nr.", contract_no
    wsM["A2"], wsM["B2"] = "Akto data", header_date
    wsM["A3"], wsM["B3"] = "Atlikimo metai", header_year_exec
    autosize(wsM)

    # BENDRAS
    ws = wb.create_sheet(SHEET_SUM)
    r = 1
    ws[f"A{r}"] = f"Bendras aktas pagal sutartį: {contract_no}"
    ws[f"A{r}"].font = Font(size=12, bold=True)
    r += 2

    subtotal_refs: List[str] = []
    for addr in addresses_selected:
        dfa = dfc[dfc[col_map["address"]].astype(str).str.strip() == addr].copy()
        meta = {
            "uzsakovas": (str(dfa.iloc[0][col_map["uzsakovas"]]).strip() if col_map["uzsakovas"] and not dfa.empty else ""),
            "vykdytojas": (str(dfa.iloc[0][col_map["vykdytojas"]]).strip() if col_map["vykdytojas"] and not dfa.empty else ""),
            "sutartis": contract_no,
            "skyrius": (str(dfa.iloc[0][col_map["skyrius"]]).strip() if col_map["skyrius"] and not dfa.empty else ""),
        }
        rows: List[Tuple[str, float, float]] = []
        for _, row in dfa.iterrows():
            rows.append((
                str(row[col_map["service"]]).strip(),
                dec2(row[col_map["qty"]]),
                dec2(row[col_map["rate"]]),
            ))
        r, subref = _write_address_block(ws, r, addr, meta, rows)
        subtotal_refs.append(subref)

    if subtotal_refs:
        ws[f"D{r}"] = "Bendra suma (be PVM):"
        ws[f"E{r}"] = f"=SUM({','.join(subtotal_refs)})"
        ws[f"E{r}"].number_format = FMT_MONEY
        r += 1

        ws[f"D{r}"] = f"PVM {float(PVM_DEFAULT)}%:"
        ws[f"E{r}"] = f"=E{r-1}*{float(PVM_DEFAULT)/100}"; ws[f"E{r}"].number_format = FMT_MONEY
        r += 1

        ws[f"D{r}"] = "Bendra suma su PVM:"
        ws[f"E{r}"] = f"=E{r-2}+E{r-1}"; ws[f"E{r}"].number_format = FMT_MONEY
        set_borders(ws, f"D{r-2}:E{r}", thick=True)

    autosize(ws)

    # informacinis KATALOGAS
    wsK = wb.create_sheet(SHEET_DB)
    wsK.append(["Adresas", "Paslauga", "Įkainis (be PVM)", "Kiekis", "Skyrius", "Užsakovas", "Vykdytojas", "Sutarties nr."])
    for _, row in dfc.iterrows():
        wsK.append([
            str(row[col_map["address"]]).strip(),
            str(row[col_map["service"]]).strip(),
            dec2(row[col_map["rate"]]),
            dec2(row[col_map["qty"]]),
            str(row[col_map["skyrius"]]).strip() if col_map["skyrius"] else "",
            str(row[col_map["uzsakovas"]]).strip() if col_map["uzsakovas"] else "",
            str(row[col_map["vykdytojas"]]).strip() if col_map["vykdytojas"] else "",
            contract_no
        ])
    autosize(wsK)

    return wb


# ===== Valymas (be tuščių eilučių) =====
def clean_before_sending_bytes(xlsx_bytes: bytes) -> bytes:
    bio = io.BytesIO(xlsx_bytes)
    wb = load_workbook(bio)
    if SHEET_AKT in wb.sheetnames:
        ws = wb[SHEET_AKT]
        start = ROW_TABLE_START
        end = ROW_TABLE_START + MAX_LINES - 1
        last_used = start - 1
        for r in range(start, min(ws.max_row, end) + 1):
            if any(ws.cell(row=r, column=c).value not in (None, "") for c in (2, 3, 4, 5)):
                last_used = r
        to_delete = end - last_used
        if to_delete > 0:
            ws.delete_rows(last_used + 1, to_delete)
        if last_used >= start:
            ws["E12"].value = f"=SUM(E{start}:E{last_used})"
        else:
            ws["E12"].value = "0"
        ws["E13"].value = f"=E12*{float(PVM_DEFAULT)/100}"
        ws["E14"].value = "=E12+E13"

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ===== STREAMLIT UI (jokio ankstyvo stop prieš UI headerius) =====
st.set_page_config(page_title="Aktų generatorius (Excel)", layout="centered")
st.title("Aktų generatorius (Excel)")
st.caption(f"Python: {sys.version}")

# Katalogo įkėlimas
with st.expander("KATALOGAS (CSV/XLSX)", expanded=True):
    up = st.file_uploader(
        "Įkelk katalogą (min.: Objekto adresas, Paslaugos pavadinimas, Įkainis, Plotas (m2); "
        "pasirenkamai: Skyrius, Užsakovas, Vykdytojas, Sutarties numeris, Vadybininkas)",
        type=["csv", "xlsx"]
    )
    if not up:
        st.info("Įkelk failą, tuomet atsiras filtrai ir mygtukai.")
        st.stop()  # <− saugu: UI jau matosi (title + uploader)

# Skaitymas + map
try:
    df_raw = read_catalog(up)
    col_map = map_columns(df_raw)
except Exception as e:
    st.error(f"Katalogo klaida: {e}")
    st.stop()

# Vadybininko filtras
df = df_raw.copy()
if col_map.get("vadybininkas"):
    with st.expander("Filtras pagal vadybininką (pasirenkama)", expanded=False):
        mgrs = sorted(set(str(x).strip() for x in df[col_map["vadybininkas"]].dropna()))
        sel_mgrs = st.multiselect("Vadybininkas(-ai)", mgrs)
        if sel_mgrs:
            df = df[df[col_map["vadybininkas"]].astype(str).str.strip().isin(sel_mgrs)]

# ——— Vieno objekto aktas ———
st.subheader("Vieno objekto aktas (pagal adresą arba skyrių)")
mode = st.radio("Pasirinkimas", ["Pagal objekto adresą", "Pagal skyrių"], horizontal=True)
selection_mode = "adresas" if mode.startswith("Pagal objekto adresą") else "skyrius"

addresses = sorted(set(str(x).strip() for x in df[col_map["address"]].dropna() if str(x).strip()))
departms = sorted(set(str(x).strip() for x in (df[col_map["skyrius"]] if col_map["skyrius"] else pd.Series(dtype=str)).dropna() if str(x).strip()))

selected_key = None
if selection_mode == "adresas":
    if addresses:
        selected_key = st.selectbox("Objekto adresas", addresses, index=0)
    else:
        st.warning("Po filtrų nerasta adresų.")
else:
    if departms:
        selected_key = st.selectbox("Skyrius", departms, index=0)
    else:
        st.warning("Po filtrų nerasta skyrių.")

c1, c2 = st.columns(2)
akto_data = c1.text_input("Akto data", "2026-01-04")
metai_atlik = c2.text_input("Atlikimo metai", "2026")

btn1 = st.button("Generuoti vieno objekto aktą (XLSX)", use_container_width=True, disabled=not bool(selected_key))
if btn1 and selected_key:
    try:
        wb_single = build_workbook_single(df, col_map, selection_mode, selected_key, akto_data, metai_atlik)
        bio = io.BytesIO(); wb_single.save(bio); xlsx = bio.getvalue()
        st.download_button("Atsisiųsti aktą", xlsx, "aktas_dropdown_paieska.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        cleaned = clean_before_sending_bytes(xlsx)
        st.download_button("Atsisiųsti švarų aktą", cleaned, "aktas_dropdown_paieska_clean.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        st.exception(e)

# ——— Bendras aktas pagal sutartį ———
st.subheader("Bendras aktas pagal sutartį (viename faile – atskiros lentelės kiekvienam adresui)")
if not col_map["sutartis"]:
    st.info("Kataloge nėra „Sutarties numeris“ – bendro akto generuoti negalėsi.")
else:
    sutartys = sorted(set(str(x).strip() for x in df[col_map["sutartis"]].dropna() if str(x).strip()))
    if sutartys:
        sel_contract = st.selectbox("Sutarties numeris", sutartys, index=0)
        dfc = df[df[col_map["sutartis"]].astype(str).str.strip() == sel_contract]
        addrs_in_contract = sorted(set(dfc[col_map["address"]].astype(str).str.strip()))
        selected_addrs = st.multiselect("Adresai (nepasirinkus – ims visus pagal sutartį)", addrs_in_contract)
        btn2 = st.button("Generuoti bendrą aktą (XLSX)", use_container_width=True)
        if btn2:
            try:
                wb_sum = build_workbook_consolidated(df, col_map, sel_contract, selected_addrs, akto_data, metai_atlik)
                bio2 = io.BytesIO(); wb_sum.save(bio2); xlsx2 = bio2.getvalue()
                st.download_button("Atsisiųsti bendrą aktą", xlsx2, "aktas_bendras_pagal_sutarti.xlsx",
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            except Exception as e:
                st.exception(e)
    else:
        st.info("Po filtrų sutarties reikšmių nerasta.")
